#!/usr/bin/env python3
"""Adaptively inspect an Excel literature list and download paper PDFs.

The script is designed for Agent invocation: workbook interpretation is confidence-
scored and auditable, low-confidence schemas stop safely, multiple source columns
are merged by priority, user-supplied sessions/headers are supported, and every
network attempt is recorded.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import ipaddress
import json
import os
import re
import shutil
import socket
import ssl
import time
import uuid
import unicodedata
from collections import defaultdict, deque
from dataclasses import dataclass, field
from email.utils import parsedate_to_datetime
from html.parser import HTMLParser
from http.cookiejar import CookieJar, MozillaCookieJar
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, quote, urljoin, urlsplit, urlunsplit
from urllib.request import HTTPCookieProcessor, HTTPRedirectHandler, HTTPSHandler, Request, build_opener

import openpyxl

from workbook_schema import (
    DEFAULT_ALIASES,
    SchemaOptions,
    dedupe_identity,
    derive_title_from_sources,
    detect_sheet_schema,
    extract_urls as _extract_urls,
    header_score,
    load_schema_config,
    merge_aliases,
    source_values_from_row,
    title_from_row,
)

USER_AGENT = "Mozilla/5.0 (compatible; DownloadPaperCollection/3.1; +https://agentskills.io/)"
SSL_CONTEXT = ssl.create_default_context()
DEFAULT_MAX_BYTES = 100 * 1024 * 1024
DEFAULT_MIN_PDF_BYTES = 1_024
DEFAULT_MAX_HTML_BYTES = 5 * 1024 * 1024
HTML_LINK_LIMIT = 40
MAX_HTML_DEPTH = 1


@dataclass
class FetchAttempt:
    url: str
    mode: str
    ok: bool
    category: str
    detail: str
    status: Optional[int] = None
    content_type: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "url": self.url,
            "mode": self.mode,
            "ok": self.ok,
            "category": self.category,
            "detail": self.detail,
            "status": self.status,
            "content_type": self.content_type,
        }


@dataclass
class FetchResult:
    final_url: str = ""
    error: str = ""
    category: str = ""
    attempts: List[FetchAttempt] = field(default_factory=list)


class DownloadError(RuntimeError):
    pass


def parse_request_headers(values: Optional[Sequence[str]]) -> Dict[str, str]:
    """Parse repeatable ``Name: Value`` CLI headers without allowing line injection."""
    headers: Dict[str, str] = {}
    for raw in values or []:
        if "\n" in raw or "\r" in raw or ":" not in raw:
            raise ValueError(f"invalid --header value: {raw!r}; expected 'Name: Value'")
        name, value = raw.split(":", 1)
        name = name.strip()
        value = value.strip()
        if not name or not re.fullmatch(r"[!#$%&'*+.^_`|~0-9A-Za-z-]+", name):
            raise ValueError(f"invalid HTTP header name: {name!r}")
        headers[name] = value
    return headers


def _load_cookie_jar(cookie_file: Optional[Path]) -> CookieJar:
    if cookie_file is None:
        return CookieJar()
    jar = MozillaCookieJar(str(cookie_file))
    try:
        jar.load(ignore_discard=True, ignore_expires=True)
    except (OSError, ValueError) as exc:
        raise ValueError(f"cannot load Netscape cookie file {cookie_file}: {exc}") from exc
    return jar


def classify_interactive_html(payload: bytes) -> str:
    """Classify common login/challenge/paywall HTML without treating it as a PDF."""
    text = payload[:524288].decode("utf-8", errors="ignore").casefold()
    captcha_markers = ("captcha", "recaptcha", "hcaptcha", "cf-chl-", "challenge-platform")
    auth_markers = (
        "sign in", "log in", "login", "institutional access", "shibboleth",
        "single sign-on", "sso", "authentication required", "登录", "统一身份认证",
    )
    paywall_markers = (
        "purchase pdf", "buy article", "subscribe to access", "access through your institution",
        "rent this article", "购买全文", "订阅后阅读",
    )
    if any(marker in text for marker in captcha_markers):
        return "interactive_challenge"
    if any(marker in text for marker in auth_markers):
        return "session_required"
    if any(marker in text for marker in paywall_markers):
        return "entitlement_required"
    return ""


class PdfLinkParser(HTMLParser):
    """Collect PDF-like links and citation metadata from one landing page."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: List[str] = []

    def _append(self, value: Optional[str]) -> None:
        if not value:
            return
        value = html.unescape(value.strip())
        if value and value not in self.links:
            self.links.append(value)

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]) -> None:
        attr = {str(key).lower(): value for key, value in attrs}
        lowered = tag.lower()
        if lowered == "meta":
            name = (attr.get("name") or attr.get("property") or attr.get("itemprop") or "").lower()
            if name in {
                "citation_pdf_url",
                "bepress_citation_pdf_url",
                "eprints.document_url",
                "wkhealth_pdf_url",
                "og:pdf",
                "pdf_url",
            }:
                self._append(attr.get("content"))
            return
        if lowered in {"iframe", "embed", "object"}:
            self._append(attr.get("src") or attr.get("data"))
            return
        if lowered not in {"a", "link", "button"}:
            return
        href = attr.get("href") or attr.get("data-href") or attr.get("data-url") or attr.get("data-pdf-url")
        if not href:
            return
        rel = (attr.get("rel") or "").lower()
        media_type = (attr.get("type") or "").lower()
        lower = href.lower()
        if (
            ".pdf" in lower
            or "application/pdf" in media_type
            or "citation_pdf" in lower
            or re.search(r"/(?:pdf|download)(?:/|$|\?)", lower)
            or "pdf=" in lower
            or ("download" in lower and "pdf" in rel)
        ):
            self._append(href)


class SafeRedirectHandler(HTTPRedirectHandler):
    def __init__(self, allow_private_hosts: bool) -> None:
        super().__init__()
        self.allow_private_hosts = allow_private_hosts

    def redirect_request(self, req, fp, code, msg, headers, newurl):  # type: ignore[override]
        error = url_safety_error(newurl, self.allow_private_hosts)
        if error:
            raise DownloadError(f"blocked redirect: {error}")
        return super().redirect_request(req, fp, code, msg, headers, newurl)


def extract_urls(
    value: object,
    hyperlink_target: Optional[str] = None,
    source_kind: str = "source",
) -> List[str]:
    """Backward-compatible re-export of the adaptive schema URL extractor."""
    return _extract_urls(value, hyperlink_target, source_kind)


def clean_name(value: object, limit: int = 180) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = re.sub(r'[/:*?"<>|\\\x00-\x1f]', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    if not text:
        text = "未命名"
    if len(text) > limit:
        digest = hashlib.sha256(text.encode("utf-8")).hexdigest()[:8]
        text = f"{text[: max(1, limit - 10)].rstrip()}__{digest}"
    return text.rstrip(" .") or "未命名"


def english_title(value: object) -> str:
    """Backward-compatible helper retained for callers and tests."""
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if not text:
        return ""
    match = re.search(
        r"\[\s*EN\s*\]\s*(.*?)(?=\[\s*(?:中文|CN|ZH)\s*\]|$)",
        text,
        flags=re.I | re.S,
    )
    candidate = match.group(1).strip() if match else text.splitlines()[0].strip()
    return re.sub(r"\s+", " ", candidate).strip()


def title_key(value: object) -> str:
    text = unicodedata.normalize("NFKC", english_title(value)).casefold()
    text = re.sub(r"[^\w]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    text = re.sub(r"[\s_\-—–:：()（）\[\]【】]+", " ", text)
    return text.strip()


def best_column(headers: Sequence[object], aliases: Sequence[str]) -> Optional[int]:
    scored = [(header_score(value, aliases), index) for index, value in enumerate(headers)]
    score, index = max(scored, default=(0, -1))
    return index if score > 0 else None


def extract_arxiv_identifier(url: str) -> str:
    parts = urlsplit(url)
    if "arxiv.org" not in (parts.hostname or "").casefold():
        return ""
    match = re.search(r"/(?:abs|pdf)/([^/?#]+)", parts.path)
    return match.group(1).removesuffix(".pdf") if match else ""


def _looks_like_direct_pdf(url: str) -> bool:
    parts = urlsplit(url)
    lower_path = parts.path.casefold()
    lower_query = parts.query.casefold()
    return lower_path.endswith(".pdf") or "format=pdf" in lower_query or "response-content-type=application%2fpdf" in lower_query


def _nonpaper_source(url: str) -> bool:
    parts = urlsplit(url)
    host = (parts.hostname or "").casefold()
    path = parts.path.casefold()
    if _looks_like_direct_pdf(url):
        return False
    if host in {"github.com", "www.github.com", "gitlab.com", "www.youtube.com", "youtube.com", "youtu.be"}:
        return True
    if any(token in path for token in ("/supplement", "/supplementary", "/code", "/dataset")):
        return True
    return False


def source_priority(url: str) -> int:
    parts = urlsplit(url)
    host = (parts.hostname or "").casefold()
    if _looks_like_direct_pdf(url):
        return 100
    if host in {
        "arxiv.org",
        "www.arxiv.org",
        "openreview.net",
        "openaccess.thecvf.com",
        "proceedings.mlr.press",
        "aclanthology.org",
        "proceedings.neurips.cc",
        "papers.nips.cc",
        "ojs.aaai.org",
    }:
        return 88
    if host in {"doi.org", "dx.doi.org"}:
        return 62
    if _nonpaper_source(url):
        return 5
    return 45


def candidate_urls(source_url: str) -> List[str]:
    if not source_url:
        return []
    url = source_url.strip()
    parts = urlsplit(url)
    host = (parts.hostname or "").casefold()
    path = parts.path
    candidates: List[str] = []

    def add(candidate: str) -> None:
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    if host in {"arxiv.org", "www.arxiv.org"}:
        identifier = extract_arxiv_identifier(url)
        if identifier:
            add(f"https://arxiv.org/pdf/{identifier}.pdf")
    if host == "openaccess.thecvf.com" and "/html/" in path:
        add(
            urlunsplit(
                parts._replace(
                    path=path.replace("/html/", "/papers/").replace("_paper.html", "_paper.pdf"),
                    query="",
                    fragment="",
                )
            )
        )
    if host in {"proceedings.neurips.cc", "papers.nips.cc"} and "Abstract-" in path:
        new_path = path.replace("Abstract-", "Paper-").replace(".html", ".pdf").replace("/html/", "/file/")
        add(urlunsplit(parts._replace(path=new_path, query="", fragment="")))
    if host == "ojs.aaai.org" and "/article/view/" in path:
        add(urlunsplit(parts._replace(path=path.replace("/view/", "/download/"), query="", fragment="")))
    if host == "openreview.net" and path.rstrip("/") == "/forum":
        paper_id = parse_qs(parts.query).get("id", [""])[0]
        if paper_id:
            add(f"https://openreview.net/pdf?id={quote(paper_id)}")
    if host == "proceedings.mlr.press" and path.endswith(".html"):
        add(urlunsplit(parts._replace(path=path[:-5] + ".pdf", query="", fragment="")))
    if host == "aclanthology.org" and not path.endswith(".pdf"):
        add(urlunsplit(parts._replace(path=path.rstrip("/") + ".pdf", query="", fragment="")))
    if host in {"hal.science", "hal.archives-ouvertes.fr"} and "/hal-" in path and not path.endswith("/document"):
        add(urlunsplit(parts._replace(path=path.rstrip("/") + "/document", query="", fragment="")))
    if not _nonpaper_source(url):
        add(url)
    return candidates


def rank_source_urls(urls: Iterable[str]) -> List[str]:
    unique: List[str] = []
    for url in urls:
        if url and url not in unique:
            unique.append(url)
    return sorted(unique, key=lambda value: (-source_priority(value), unique.index(value)))


SOURCE_SEMANTIC_BOOST = {
    "pdf_url": 35,
    "arxiv": 22,
    "openreview": 22,
    "doi": 12,
    "source": 0,
}


def rank_source_details(details: Sequence[Mapping[str, Any]]) -> List[str]:
    scored: Dict[str, Tuple[int, int]] = {}
    order = 0
    for detail in details:
        semantic = str(detail.get("semantic") or "source")
        boost = SOURCE_SEMANTIC_BOOST.get(semantic, 0)
        urls = detail.get("urls") or []
        if not isinstance(urls, list):
            continue
        for raw_url in urls:
            url = str(raw_url).strip()
            if not url:
                continue
            score = source_priority(url) + boost
            current = scored.get(url)
            if current is None or score > current[0]:
                scored[url] = (score, order)
            order += 1
    return [url for url, _ in sorted(scored.items(), key=lambda item: (-item[1][0], item[1][1]))]


def html_pdf_links(final_url: str, content: bytes) -> List[str]:
    parser = PdfLinkParser()
    parser.feed(content.decode("utf-8", errors="ignore"))
    links: List[str] = []
    for link in parser.links[:HTML_LINK_LIMIT]:
        absolute = urljoin(final_url, link)
        if absolute not in links:
            links.append(absolute)
    return sorted(links, key=lambda value: -source_priority(value))


def valid_pdf(path: Path, min_bytes: int = DEFAULT_MIN_PDF_BYTES) -> bool:
    if not path.is_file() or path.stat().st_size < min_bytes:
        return False
    with path.open("rb") as handle:
        head = handle.read(8)
        if not head.startswith(b"%PDF-"):
            return False
        tail_size = min(path.stat().st_size, 8192)
        handle.seek(-tail_size, os.SEEK_END)
        tail = handle.read()
    return b"%%EOF" in tail


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def url_safety_error(url: str, allow_private_hosts: bool = False) -> str:
    parts = urlsplit(url)
    if parts.scheme not in {"http", "https"}:
        return f"unsupported URL scheme {parts.scheme!r}: {url}"
    hostname = parts.hostname
    if not hostname:
        return f"missing hostname: {url}"
    if allow_private_hosts:
        return ""
    if hostname.casefold() == "localhost" or hostname.casefold().endswith(".local"):
        return f"private hostname is blocked: {hostname}"
    try:
        addresses = {
            info[4][0]
            for info in socket.getaddrinfo(
                hostname,
                parts.port or (443 if parts.scheme == "https" else 80),
                type=socket.SOCK_STREAM,
            )
        }
    except socket.gaierror as exc:
        return f"DNS resolution failed for {hostname}: {exc}"
    for address in addresses:
        ip = ipaddress.ip_address(address)
        if (
            ip.is_private
            or ip.is_loopback
            or ip.is_link_local
            or ip.is_multicast
            or ip.is_reserved
            or ip.is_unspecified
        ):
            return f"private or non-routable address is blocked: {hostname} -> {ip}"
    return ""


def _stream_response(response: Any, destination: Path, max_bytes: int) -> int:
    content_length = response.headers.get("content-length")
    if content_length:
        try:
            if int(content_length) > max_bytes:
                raise DownloadError(f"response exceeds --max-bytes ({content_length} > {max_bytes})")
        except ValueError:
            pass
    total = 0
    with destination.open("wb") as handle:
        while True:
            chunk = response.read(64 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > max_bytes:
                raise DownloadError(f"response exceeds --max-bytes ({max_bytes})")
            handle.write(chunk)
    return total


def _retry_after_seconds(headers: Mapping[str, str], fallback: float) -> float:
    raw = headers.get("Retry-After") or headers.get("retry-after")
    if not raw:
        return fallback
    try:
        return min(30.0, max(0.0, float(raw)))
    except ValueError:
        try:
            dt = parsedate_to_datetime(raw)
            return min(30.0, max(0.0, dt.timestamp() - time.time()))
        except (TypeError, ValueError, OverflowError):
            return fallback


def _classify_http_status(status: int) -> str:
    if status == 401:
        return "authorization_required"
    if status == 403:
        return "forbidden_or_challenge"
    if status == 404:
        return "not_found"
    if status == 429:
        return "rate_limited"
    if status >= 500:
        return "server_error"
    return "http_error"


def fetch_pdf_detailed(
    urls: Sequence[str],
    destination: Path,
    timeout: int,
    retries: int,
    max_bytes: int,
    min_pdf_bytes: int,
    allow_private_hosts: bool,
    host_delay: float = 0.2,
    retry_delay: float = 1.0,
    max_html_bytes: int = DEFAULT_MAX_HTML_BYTES,
    request_headers: Optional[Mapping[str, str]] = None,
    cookie_header: str = "",
    cookie_file: Optional[Path] = None,
    user_agent: str = USER_AGENT,
) -> FetchResult:
    result = FetchResult()
    ranked = rank_source_urls(urls)
    if not ranked:
        result.error = "no source URL"
        result.category = "missing_source"
        return result

    queue: deque[Tuple[str, int, str, str]] = deque()
    for source in ranked:
        for candidate in candidate_urls(source):
            queue.append((candidate, 0, "direct-transform" if candidate != source else "source", ""))
    seen: set[str] = set()
    last_host_access: Dict[str, float] = {}
    try:
        cookie_jar = _load_cookie_jar(cookie_file)
    except ValueError as exc:
        result.error = str(exc)
        result.category = "session_configuration"
        return result
    opener = build_opener(
        HTTPCookieProcessor(cookie_jar),
        SafeRedirectHandler(allow_private_hosts),
        HTTPSHandler(context=SSL_CONTEXT),
    )
    part_path = destination.with_suffix(destination.suffix + ".part")
    destination.parent.mkdir(parents=True, exist_ok=True)

    while queue:
        url, depth, mode, referer = queue.popleft()
        if url in seen:
            continue
        seen.add(url)
        safety_error = url_safety_error(url, allow_private_hosts)
        if safety_error:
            result.attempts.append(FetchAttempt(url, mode, False, "blocked_url", safety_error))
            continue

        host = (urlsplit(url).hostname or "").casefold()
        elapsed = time.monotonic() - last_host_access.get(host, 0.0)
        if host_delay > 0 and elapsed < host_delay:
            time.sleep(host_delay - elapsed)

        for attempt_number in range(retries + 1):
            part_path.unlink(missing_ok=True)
            effective_headers = {
                "User-Agent": user_agent,
                "Accept": "application/pdf,text/html;q=0.9,application/xhtml+xml;q=0.8,*/*;q=0.5",
                "Accept-Language": "en-US,en;q=0.8,zh-CN;q=0.6",
            }
            effective_headers.update(dict(request_headers or {}))
            if cookie_header:
                existing_cookie = effective_headers.get("Cookie", "")
                effective_headers["Cookie"] = "; ".join(value for value in (existing_cookie, cookie_header) if value)
            if referer and "Referer" not in effective_headers:
                effective_headers["Referer"] = referer
            request = Request(url, headers=effective_headers)
            try:
                last_host_access[host] = time.monotonic()
                with opener.open(request, timeout=timeout) as response:
                    final_url = response.geturl()
                    redirect_error = url_safety_error(final_url, allow_private_hosts)
                    if redirect_error:
                        raise DownloadError(f"blocked final URL: {redirect_error}")
                    content_type = (response.headers.get("content-type") or "").split(";", 1)[0].strip().casefold()
                    _stream_response(response, part_path, max_bytes)
                    if valid_pdf(part_path, min_pdf_bytes):
                        part_path.replace(destination)
                        result.final_url = final_url
                        result.category = "success"
                        result.attempts.append(
                            FetchAttempt(final_url, mode, True, "success", "valid PDF", getattr(response, "status", 200), content_type)
                        )
                        return result

                    size = part_path.stat().st_size if part_path.exists() else 0
                    page = b""
                    if part_path.exists() and size <= max_html_bytes:
                        page = part_path.read_bytes()
                    page_category = classify_interactive_html(page) if page else ""
                    category = page_category or "non_pdf"
                    detail = f"response is not a valid PDF ({size} bytes)"
                    if page_category:
                        detail += f"; detected {page_category.replace('_', ' ')} page"
                    result.attempts.append(
                        FetchAttempt(final_url, mode, False, category, detail, getattr(response, "status", 200), content_type)
                    )
                    if depth < MAX_HTML_DEPTH and page:
                        for link in html_pdf_links(final_url, page):
                            if link not in seen:
                                queue.appendleft((link, depth + 1, "html-discovery", final_url))
                    part_path.unlink(missing_ok=True)
                    break
            except HTTPError as exc:
                category = _classify_http_status(exc.code)
                detail = f"HTTP {exc.code}: {exc.reason}"
                result.attempts.append(FetchAttempt(url, mode, False, category, detail, exc.code))
                if exc.code in {429} or exc.code >= 500:
                    if attempt_number < retries:
                        time.sleep(_retry_after_seconds(exc.headers, retry_delay * (attempt_number + 1)))
                        continue
                break
            except (TimeoutError, socket.timeout) as exc:
                result.attempts.append(FetchAttempt(url, mode, False, "timeout", str(exc) or "request timed out"))
                if attempt_number < retries:
                    time.sleep(retry_delay * (attempt_number + 1))
                    continue
                break
            except (URLError, OSError, DownloadError) as exc:
                category = "network_error"
                if isinstance(exc, DownloadError) and "blocked" in str(exc).casefold():
                    category = "blocked_url"
                result.attempts.append(FetchAttempt(url, mode, False, category, str(exc)))
                if attempt_number < retries and category == "network_error":
                    time.sleep(retry_delay * (attempt_number + 1))
                    continue
                break
            finally:
                if part_path.exists() and not valid_pdf(part_path, min_pdf_bytes):
                    part_path.unlink(missing_ok=True)

    categories = [attempt.category for attempt in result.attempts]
    priority = [
        "interactive_challenge",
        "session_required",
        "entitlement_required",
        "authorization_required",
        "forbidden_or_challenge",
        "rate_limited",
        "not_found",
        "timeout",
        "server_error",
        "non_pdf",
        "blocked_url",
        "network_error",
        "http_error",
    ]
    result.category = next((category for category in priority if category in categories), "download_failed")
    messages = [f"{attempt.url}: {attempt.detail}" for attempt in result.attempts[-8:]]
    result.error = " | ".join(messages) or "all source candidates failed"
    return result


def fetch_pdf(
    urls: Sequence[str],
    destination: Path,
    timeout: int,
    retries: int,
    max_bytes: int,
    min_pdf_bytes: int,
    allow_private_hosts: bool,
) -> Tuple[str, str]:
    """Backward-compatible wrapper returning final URL and error only."""
    result = fetch_pdf_detailed(
        urls,
        destination,
        timeout,
        retries,
        max_bytes,
        min_pdf_bytes,
        allow_private_hosts,
        host_delay=0,
    )
    return result.final_url, result.error


def _row_is_repeated_header(sheet: Any, row: int, schema: Any) -> bool:
    title_cells = [sheet.cell(row=row, column=binding.index).value for binding in schema.title_columns]
    source_cells = [sheet.cell(row=row, column=binding.index).value for binding in schema.source_columns]
    title_aliases = DEFAULT_ALIASES["title"] + DEFAULT_ALIASES["title_en"] + DEFAULT_ALIASES["title_zh"]
    source_aliases = sum((DEFAULT_ALIASES[key] for key in ("pdf_url", "doi", "arxiv", "openreview", "source")), [])
    return any(header_score(value, title_aliases) >= 80 for value in title_cells) and any(
        header_score(value, source_aliases) >= 75 for value in source_cells
    )


def inspect_workbook(
    workbook_path: Path,
    selected_sheets: Optional[Sequence[str]] = None,
    header_row: Optional[int] = None,
    *,
    schema_config: Optional[Mapping[str, Any]] = None,
    min_schema_confidence: float = 0.72,
    include_hidden_sheets: bool = False,
    strict_schema: bool = False,
    title_column: Optional[str] = None,
    venue_column: Optional[str] = None,
    source_columns: Optional[Sequence[str]] = None,
) -> Tuple[Dict[str, object], List[Dict[str, object]]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
    selected = set(selected_sheets or workbook.sheetnames)
    unknown = sorted(selected - set(workbook.sheetnames))
    if unknown:
        workbook.close()
        raise ValueError(f"Unknown worksheets: {unknown}; available: {workbook.sheetnames}")

    aliases = merge_aliases(schema_config)
    options = SchemaOptions(
        header_row=header_row,
        min_confidence=min_schema_confidence,
        include_hidden_sheets=include_hidden_sheets,
        strict_schema=strict_schema,
        title_column=title_column,
        venue_column=venue_column,
        source_columns=list(source_columns or []),
        schema_config=schema_config,
    )

    report_sheets: List[Dict[str, object]] = []
    items: List[Dict[str, object]] = []
    issues: List[Dict[str, object]] = []
    explicitly_selected = selected_sheets is not None

    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name not in selected:
                continue
            sheet = workbook[sheet_name]
            try:
                schema = detect_sheet_schema(sheet, options, aliases)
            except ValueError as exc:
                severity = "error"
                issue = {"sheet": sheet_name, "severity": severity, "category": "schema_config", "message": str(exc)}
                issues.append(issue)
                report_sheets.append({"sheet": sheet_name, "status": "error", "records": 0, "error": str(exc)})
                continue

            schema_report = schema.to_dict()
            schema_report["records"] = 0
            schema_report["issues"] = []

            if schema.status != "accepted":
                if schema.status == "empty":
                    severity = "info"
                elif schema.status == "hidden" and explicitly_selected:
                    severity = "error"
                elif schema.status == "hidden":
                    severity = "info"
                elif explicitly_selected or strict_schema:
                    severity = "error"
                else:
                    severity = "info"
                message = (
                    f"sheet skipped: schema status={schema.status}, confidence={schema.confidence:.3f}; "
                    + "; ".join(schema.reasons)
                )
                issue = {"sheet": sheet_name, "severity": severity, "category": "schema", "message": message}
                issues.append(issue)
                schema_report["issues"].append(issue)
                report_sheets.append(schema_report)
                continue

            for warning in schema.warnings:
                issue = {
                    "sheet": sheet_name,
                    "severity": "warning",
                    "category": "schema_warning",
                    "message": warning,
                }
                issues.append(issue)
                schema_report["issues"].append(issue)

            assert schema.header_row is not None
            blank_streak = 0
            sheet_records = 0
            for row_number in range(schema.header_row + 1, sheet.max_row + 1):
                if _row_is_repeated_header(sheet, row_number, schema):
                    continue
                title, title_en, title_zh = title_from_row(sheet, row_number, schema)
                urls, source_details = source_values_from_row(sheet, row_number, schema)
                ranked_urls = rank_source_details(source_details) or rank_source_urls(urls)
                venue = "未分类"
                if schema.venue_column:
                    venue = str(sheet.cell(row=row_number, column=schema.venue_column.index).value or "未分类").strip() or "未分类"
                year = ""
                if schema.year_column:
                    year = str(sheet.cell(row=row_number, column=schema.year_column.index).value or "").strip()

                if not title and not ranked_urls and venue == "未分类" and not year:
                    blank_streak += 1
                    if blank_streak >= 100:
                        break
                    continue
                blank_streak = 0

                row_issues: List[str] = []
                derived_title = False
                if not title and ranked_urls:
                    title = derive_title_from_sources(ranked_urls, row_number)
                    derived_title = True
                    row_issues.append("title missing; deterministic title derived from DOI/URL")
                if not title:
                    issue = {
                        "sheet": sheet_name,
                        "row": row_number,
                        "severity": "warning",
                        "category": "missing_title",
                        "message": "row skipped because both title and downloadable source identity are missing",
                    }
                    issues.append(issue)
                    schema_report["issues"].append(issue)
                    continue
                if not ranked_urls:
                    row_issues.append("no HTTP(S) URL, DOI, arXiv ID, or OpenReview ID detected")
                for message in row_issues:
                    issue = {
                        "sheet": sheet_name,
                        "row": row_number,
                        "severity": "warning",
                        "category": "row_data",
                        "message": message,
                    }
                    issues.append(issue)
                    schema_report["issues"].append(issue)

                dedupe_key, dedupe_mode = dedupe_identity(title, ranked_urls)
                item = {
                    "record_id": f"{sheet_name}!{row_number}",
                    "sheet": sheet_name,
                    "row": row_number,
                    "title": title,
                    "title_en": title_en,
                    "title_zh": title_zh,
                    "title_derived": derived_title,
                    "title_key": title_key(title),
                    "dedupe_key": dedupe_key,
                    "dedupe_mode": dedupe_mode,
                    "venue": venue,
                    "year": year,
                    "source_urls": ranked_urls,
                    "source_details": source_details,
                    "schema_confidence": schema.confidence,
                }
                items.append(item)
                sheet_records += 1

            schema_report["records"] = sheet_records
            report_sheets.append(schema_report)
    finally:
        workbook.close()

    accepted_sheets = [sheet for sheet in report_sheets if sheet.get("status") == "accepted"]
    if not accepted_sheets:
        issues.append(
            {
                "severity": "error",
                "category": "schema",
                "message": "no worksheet passed schema confidence; use --schema-config or explicit column options",
            }
        )
    report: Dict[str, object] = {
        "workbook": str(workbook_path),
        "schema_version": "3.1",
        "min_schema_confidence": min_schema_confidence,
        "worksheets": report_sheets,
        "accepted_sheets": [sheet["sheet"] for sheet in accepted_sheets],
        "records": len(items),
        "unique_papers": len({item["dedupe_key"] for item in items}),
        "unique_titles": len({item["title_key"] for item in items}),
        "issues": issues,
        "ok": not any(issue.get("severity") == "error" for issue in issues),
    }
    return report, items


def assign_targets(items: List[Dict[str, object]], output: Path) -> None:
    used: Dict[str, str] = {}
    for item in items:
        directory = output / clean_name(item["sheet"], 100) / clean_name(item["venue"], 120)
        base = clean_name(item["title"], 180)
        candidate = directory / f"{base}.pdf"
        identity = str(item.get("dedupe_key") or item.get("title_key") or title_key(item.get("title", "")))
        relative_key = candidate.relative_to(output).as_posix().casefold()
        if relative_key in used and used[relative_key] != identity:
            suffix = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:8]
            candidate = directory / f"{clean_name(base, 169)}__{suffix}.pdf"
            relative_key = candidate.relative_to(output).as_posix().casefold()
        used[relative_key] = identity
        item["target"] = str(candidate)


def atomic_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(text, encoding=encoding)
    temp.replace(path)


def append_jsonl(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(dict(payload), ensure_ascii=False) + "\n")
        handle.flush()
        os.fsync(handle.fileno())


def write_manifests(output: Path, rows: List[Dict[str, object]]) -> Tuple[Path, Path]:
    csv_path = output / "下载清单.csv"
    json_path = output / "下载清单.json"
    fieldnames = [
        "record_id",
        "sheet",
        "row",
        "title",
        "title_en",
        "title_zh",
        "title_derived",
        "venue",
        "year",
        "dedupe_key",
        "dedupe_mode",
        "schema_confidence",
        "source_url",
        "source_columns",
        "status",
        "resolution",
        "failure_category",
        "pdf_url",
        "local_path",
        "bytes",
        "sha256",
        "attempts",
        "error",
    ]
    temp_csv = csv_path.with_suffix(".csv.tmp")
    with temp_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    temp_csv.replace(csv_path)
    atomic_write_text(json_path, json.dumps(rows, ensure_ascii=False, indent=2) + "\n")
    return csv_path, json_path


def copy_workbook(input_path: Path, output: Path) -> Optional[Path]:
    destination = output / input_path.name
    try:
        if input_path.resolve() == destination.resolve():
            return destination
    except FileNotFoundError:
        pass
    shutil.copy2(input_path, destination)
    return destination


def _load_prior_manifest(output: Path) -> Dict[str, Path]:
    path = output / "下载清单.json"
    if not path.is_file():
        return {}
    try:
        rows = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    result: Dict[str, Path] = {}
    if isinstance(rows, list):
        for row in rows:
            if not isinstance(row, dict) or row.get("status") != "成功":
                continue
            local_path = row.get("local_path")
            dedupe_key = row.get("dedupe_key")
            if isinstance(local_path, str) and isinstance(dedupe_key, str):
                candidate = output / local_path
                if valid_pdf(candidate):
                    result[dedupe_key] = candidate
    return result


def _quarantine_invalid(target: Path) -> None:
    if not target.exists() or valid_pdf(target):
        return
    quarantine = target.with_suffix(target.suffix + ".invalid")
    counter = 1
    while quarantine.exists():
        quarantine = target.with_suffix(target.suffix + f".invalid-{counter}")
        counter += 1
    target.replace(quarantine)


def _copy_verified_pdf(source: Path, target: Path, min_pdf_bytes: int) -> bool:
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists() and not valid_pdf(target, min_pdf_bytes):
        _quarantine_invalid(target)
    if source.resolve() != target.resolve():
        temp_target = target.with_suffix(target.suffix + ".tmp")
        shutil.copy2(source, temp_target)
        temp_target.replace(target)
    return valid_pdf(target, min_pdf_bytes)


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Adaptively inspect an Excel literature list and download paper PDFs.")
    parser.add_argument("--input", required=True, type=Path, help=".xlsx or .xlsm workbook")
    parser.add_argument("--output", type=Path, help="collection directory; required unless --inspect-only")
    parser.add_argument("--inspect-only", "--dry-run", action="store_true", help="detect schemas and print a JSON preflight report")
    parser.add_argument("--header-row", type=int, help="global header-row override")
    parser.add_argument("--sheets", nargs="+", help="only process these worksheets")
    parser.add_argument("--schema-config", type=Path, help="JSON aliases and global/per-sheet explicit column mappings")
    parser.add_argument("--schema-out", type=Path, help="write the preflight schema report to this JSON path")
    parser.add_argument("--min-schema-confidence", type=float, default=0.72)
    parser.add_argument("--strict-schema", action="store_true", help="treat every nonempty unrecognized selected sheet as an error")
    parser.add_argument("--include-hidden-sheets", action="store_true")
    parser.add_argument("--title-column", help="explicit title column (name, letter, or 1-based index)")
    parser.add_argument("--venue-column", help="explicit venue column (name, letter, or 1-based index)")
    parser.add_argument("--source-columns", nargs="+", help="explicit source/DOI/PDF columns")
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--retries", type=int, default=2)
    parser.add_argument("--retry-delay", type=float, default=1.0)
    parser.add_argument("--host-delay", type=float, default=0.2)
    parser.add_argument("--max-bytes", type=int, default=DEFAULT_MAX_BYTES)
    parser.add_argument("--max-html-bytes", type=int, default=DEFAULT_MAX_HTML_BYTES)
    parser.add_argument("--min-pdf-bytes", type=int, default=DEFAULT_MIN_PDF_BYTES)
    parser.add_argument("--resume", action="store_true", help="reuse verified files from targets or a prior manifest")
    parser.add_argument("--force", action="store_true", help="redownload while preserving a valid fallback if download fails")
    parser.add_argument("--allow-private-hosts", action="store_true", help="allow localhost/private IP URLs only for controlled tests")
    parser.add_argument("--header", action="append", default=[], help="repeatable request header in 'Name: Value' form")
    parser.add_argument("--authorization", help="Authorization header value, for example 'Bearer ...' or 'Basic ...'")
    parser.add_argument("--cookie", default="", help="Cookie header value from a user-authorized browser/session")
    parser.add_argument("--cookie-file", type=Path, help="Netscape/Mozilla cookies.txt file exported from a user-authorized session")
    parser.add_argument("--user-agent", default=USER_AGENT, help="override the HTTP User-Agent")
    parser.add_argument("--fail-on-warning", action="store_true")
    return parser.parse_args(argv)


def _validate_args(args: argparse.Namespace) -> None:
    if not args.input.is_file():
        raise SystemExit(f"Input workbook not found: {args.input}")
    if args.input.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise SystemExit("Input must be an .xlsx or .xlsm workbook")
    if args.header_row is not None and args.header_row < 1:
        raise SystemExit("--header-row must be >= 1")
    if not 0 <= args.min_schema_confidence <= 1:
        raise SystemExit("--min-schema-confidence must be between 0 and 1")
    if args.timeout < 1 or args.retries < 0 or args.max_bytes < 1 or args.min_pdf_bytes < 1:
        raise SystemExit("timeout/max-byte values must be positive and retries must be >= 0")
    if args.host_delay < 0 or args.retry_delay < 0:
        raise SystemExit("--host-delay and --retry-delay must be >= 0")
    if args.cookie_file is not None and not args.cookie_file.is_file():
        raise SystemExit(f"Cookie file not found: {args.cookie_file}")
    try:
        parse_request_headers(args.header)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc


def run(args: argparse.Namespace) -> Dict[str, object]:
    _validate_args(args)
    try:
        schema_config = load_schema_config(args.schema_config)
        report, items = inspect_workbook(
            args.input,
            args.sheets,
            args.header_row,
            schema_config=schema_config,
            min_schema_confidence=args.min_schema_confidence,
            include_hidden_sheets=args.include_hidden_sheets,
            strict_schema=args.strict_schema,
            title_column=args.title_column,
            venue_column=args.venue_column,
            source_columns=args.source_columns,
        )
    except (ValueError, OSError, openpyxl.utils.exceptions.InvalidFileException) as exc:
        raise SystemExit(f"Workbook preflight failed: {exc}") from exc

    print(json.dumps({"event": "preflight", **report}, ensure_ascii=False), flush=True)
    if args.schema_out:
        atomic_write_text(args.schema_out, json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    warnings = [issue for issue in report["issues"] if issue.get("severity") == "warning"]  # type: ignore[index]
    errors = [issue for issue in report["issues"] if issue.get("severity") == "error"]  # type: ignore[index]
    if args.inspect_only:
        if errors or (args.fail_on_warning and warnings):
            raise SystemExit(2)
        return report
    if errors:
        raise SystemExit("Preflight failed; inspect candidate columns or provide --schema-config / explicit column options")
    if args.fail_on_warning and warnings:
        raise SystemExit("Preflight produced warnings and --fail-on-warning is set")
    if args.output is None:
        raise SystemExit("--output is required unless --inspect-only is used")

    output = args.output
    output.mkdir(parents=True, exist_ok=True)
    extra_headers = parse_request_headers(args.header)
    if args.authorization:
        extra_headers["Authorization"] = args.authorization
    atomic_write_text(output / "结构识别报告.json", json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    assign_targets(items, output)

    grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for item in items:
        grouped[str(item["dedupe_key"])].append(item)

    prior_manifest = _load_prior_manifest(output) if args.resume else {}
    journal = output / "运行日志.jsonl"
    run_id = f"{time.strftime('%Y%m%dT%H%M%SZ', time.gmtime())}-{uuid.uuid4().hex[:8]}"
    append_jsonl(
        journal,
        {
            "event": "run_start",
            "run_id": run_id,
            "input": str(args.input),
            "accepted_sheets": report["accepted_sheets"],
            "records": len(items),
            "unique_papers": len(grouped),
            "resume": bool(args.resume),
            "force": bool(args.force),
            "request_header_names": sorted(extra_headers),
            "cookie_supplied": bool(args.cookie or args.cookie_file),
        },
    )
    cache_dir = output / ".download-cache"
    cache_dir.mkdir(exist_ok=True)
    rows: List[Dict[str, object]] = []
    group_success: Dict[str, bool] = {}

    for index, (key, group) in enumerate(grouped.items(), 1):
        representative = group[0]
        planned_targets = [Path(item["target"]) for item in group]
        fallback_existing = next((target for target in planned_targets if valid_pdf(target, args.min_pdf_bytes)), None)
        if fallback_existing is None and args.resume:
            prior = prior_manifest.get(key)
            if prior and valid_pdf(prior, args.min_pdf_bytes):
                fallback_existing = prior

        if fallback_existing is not None and not args.force:
            fetch_result = FetchResult(final_url="", error="", category="success")
            source_path: Optional[Path] = fallback_existing
            resolution = "reused"
        else:
            all_source_details: List[Mapping[str, Any]] = []
            all_urls: List[str] = []
            for item in group:
                details = item.get("source_details") or []
                if isinstance(details, list):
                    all_source_details.extend(detail for detail in details if isinstance(detail, Mapping))
                for source in item["source_urls"]:  # type: ignore[index]
                    if source not in all_urls:
                        all_urls.append(str(source))
            ranked_group_urls = rank_source_details(all_source_details) or rank_source_urls(all_urls)
            cache_path = cache_dir / f"{hashlib.sha256(key.encode('utf-8')).hexdigest()[:24]}.pdf"
            if args.resume and valid_pdf(cache_path, args.min_pdf_bytes):
                fetch_result = FetchResult(final_url="", error="", category="success")
                source_path = cache_path
                resolution = "recovered-cache"
            else:
                cache_path.unlink(missing_ok=True)
                fetch_result = fetch_pdf_detailed(
                    ranked_group_urls,
                    cache_path,
                    args.timeout,
                    args.retries,
                    args.max_bytes,
                    args.min_pdf_bytes,
                    args.allow_private_hosts,
                    args.host_delay,
                    args.retry_delay,
                    args.max_html_bytes,
                    request_headers=extra_headers,
                    cookie_header=args.cookie,
                    cookie_file=args.cookie_file,
                    user_agent=args.user_agent,
                )
                if valid_pdf(cache_path, args.min_pdf_bytes):
                    source_path = cache_path
                    resolution = "downloaded"
                elif fallback_existing is not None:
                    source_path = fallback_existing
                    resolution = "reused-after-download-failure"
                else:
                    source_path = None
                    resolution = "failed"

        finalized_targets: Dict[str, bool] = {}
        if source_path:
            for target in planned_targets:
                finalized_targets[str(target)] = _copy_verified_pdf(source_path, target, args.min_pdf_bytes)
        group_ok = bool(finalized_targets) and all(finalized_targets.values())
        group_success[key] = group_ok

        for item in group:
            target = Path(item["target"])
            status = "成功" if group_ok and valid_pdf(target, args.min_pdf_bytes) else "失败"
            local_path = target.relative_to(output).as_posix() if status == "成功" else ""
            attempts = [attempt.to_dict() for attempt in fetch_result.attempts]
            error = fetch_result.error
            failure_category = "" if status == "成功" else (fetch_result.category or "final_validation")
            if status != "成功" and source_path and not valid_pdf(target, args.min_pdf_bytes):
                error = (error + " | " if error else "") + "copied file failed final PDF validation"
            rows.append(
                {
                    "record_id": item["record_id"],
                    "sheet": item["sheet"],
                    "row": item["row"],
                    "title": item["title"],
                    "title_en": item["title_en"],
                    "title_zh": item["title_zh"],
                    "title_derived": item["title_derived"],
                    "venue": item["venue"],
                    "year": item["year"],
                    "dedupe_key": item["dedupe_key"],
                    "dedupe_mode": item["dedupe_mode"],
                    "schema_confidence": round(float(item["schema_confidence"]), 3),
                    "source_url": "\n".join(item["source_urls"]),  # type: ignore[arg-type]
                    "source_columns": json.dumps(item["source_details"], ensure_ascii=False, separators=(",", ":")),
                    "status": status,
                    "resolution": resolution,
                    "failure_category": failure_category,
                    "pdf_url": fetch_result.final_url,
                    "local_path": local_path,
                    "bytes": target.stat().st_size if status == "成功" else "",
                    "sha256": sha256_file(target) if status == "成功" else "",
                    "attempts": json.dumps(attempts, ensure_ascii=False, separators=(",", ":")),
                    "error": error,
                }
            )

        append_jsonl(
            journal,
            {
                "event": "group_complete",
                "run_id": run_id,
                "progress": f"{index}/{len(grouped)}",
                "dedupe_key": key,
                "title": representative["title"],
                "ok": group_ok,
                "resolution": resolution,
                "failure_category": "" if group_ok else fetch_result.category,
                "attempts": [attempt.to_dict() for attempt in fetch_result.attempts],
            },
        )
        print(
            json.dumps(
                {
                    "event": "progress",
                    "progress": f"{index}/{len(grouped)}",
                    "title": representative["title"],
                    "ok": group_ok,
                    "resolution": resolution,
                    "failure_category": "" if group_ok else fetch_result.category,
                },
                ensure_ascii=False,
            ),
            flush=True,
        )

    csv_manifest, json_manifest = write_manifests(output, rows)
    copy_workbook(args.input, output)
    shutil.rmtree(cache_dir, ignore_errors=True)

    failure_categories: Dict[str, int] = defaultdict(int)
    for row in rows:
        if row["status"] != "成功":
            failure_categories[str(row["failure_category"] or "unknown")] += 1
    summary: Dict[str, object] = {
        "event": "summary",
        "schema_version": "3.1",
        "accepted_sheets": report["accepted_sheets"],
        "records": len(rows),
        "unique_papers": len(grouped),
        "record_success": sum(row["status"] == "成功" for row in rows),
        "record_failed": sum(row["status"] != "成功" for row in rows),
        "unique_success": sum(group_success.values()),
        "unique_failed": len(grouped) - sum(group_success.values()),
        "failure_categories": dict(sorted(failure_categories.items())),
        "manifest_csv": str(csv_manifest),
        "manifest_json": str(json_manifest),
        "schema_report": str(output / "结构识别报告.json"),
        "run_journal": str(journal),
        "output": str(output),
    }
    summary["run_id"] = run_id
    atomic_write_text(output / "下载报告.json", json.dumps(summary, ensure_ascii=False, indent=2) + "\n")
    append_jsonl(journal, {"event": "run_end", "run_id": run_id, **summary})
    print(json.dumps(summary, ensure_ascii=False), flush=True)
    return summary


def main(argv: Optional[Sequence[str]] = None) -> None:
    run(parse_args(argv))


if __name__ == "__main__":
    main()
