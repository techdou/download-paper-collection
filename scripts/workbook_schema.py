#!/usr/bin/env python3
"""Adaptive workbook schema detection for literature-list spreadsheets.

The module deliberately separates deterministic Excel interpretation from network
work. It supports multilingual aliases, merged/multi-row headers, multiple source
columns, explicit per-sheet mappings, confidence scoring, and safe low-confidence
stops.
"""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from urllib.parse import urlsplit

from openpyxl.utils import column_index_from_string

HEADER_SCAN_LIMIT = 40
MAX_HEADER_SPAN = 3
PROFILE_ROWS = 30

DEFAULT_ALIASES: Dict[str, List[str]] = {
    "title": [
        "title",
        "paper title",
        "article title",
        "publication title",
        "论文标题",
        "论文题目",
        "文献标题",
        "文献名称",
        "文章标题",
        "题名",
        "题目",
        "标题",
        "名称",
    ],
    "title_en": [
        "english title",
        "title en",
        "title english",
        "英文标题",
        "英文题目",
        "英文题名",
        "论文英文标题",
        "title [en]",
    ],
    "title_zh": [
        "chinese title",
        "title zh",
        "title cn",
        "中文标题",
        "中文题目",
        "中文题名",
        "论文中文标题",
        "title [中文]",
    ],
    "venue": [
        "venue",
        "publication venue",
        "journal/conference",
        "journal or conference",
        "journal",
        "conference",
        "期刊/会议",
        "会议/期刊",
        "发表期刊",
        "发表会议",
        "期刊名称",
        "会议名称",
        "期刊",
        "会议",
        "出处",
        "来源刊物",
        "发表平台",
    ],
    "pdf_url": [
        "pdf url",
        "pdf link",
        "full text pdf",
        "download url",
        "download link",
        "pdf链接",
        "pdf地址",
        "pdf下载",
        "全文链接",
        "全文地址",
        "下载链接",
        "下载地址",
    ],
    "doi": [
        "doi",
        "doi url",
        "doi link",
        "doi号",
        "doi链接",
        "数字对象标识符",
    ],
    "arxiv": [
        "arxiv",
        "arxiv id",
        "arxiv url",
        "预印本",
        "预印本链接",
    ],
    "openreview": [
        "openreview",
        "openreview url",
        "openreview link",
        "openreview链接",
    ],
    "source": [
        "url",
        "paper url",
        "article url",
        "publication url",
        "source url",
        "paper link",
        "article link",
        "source link",
        "web link",
        "链接",
        "网址",
        "论文链接",
        "论文地址",
        "文献链接",
        "文献地址",
        "原文链接",
        "原文地址",
        "来源链接",
        "来源地址",
        "公开链接",
        "访问地址",
    ],
    "year": [
        "year",
        "publication year",
        "published year",
        "年份",
        "发表年份",
        "出版年份",
    ],
}

SOURCE_SEMANTICS = ("pdf_url", "doi", "arxiv", "openreview", "source")
TITLE_SEMANTICS = ("title_en", "title", "title_zh")

DOI_RE = re.compile(r"(?<![\w.])(10\.\d{4,9}/[-._;()/:A-Z0-9]+)", re.I)
URL_RE = re.compile(r"https?://[^\s<>\]\[\"']+", re.I)
ARXIV_ID_RE = re.compile(r"(?:arxiv\s*:\s*)?(\d{4}\.\d{4,5}(?:v\d+)?)", re.I)
HYPERLINK_RE = re.compile(r"HYPERLINK\(\s*[\"']([^\"']+)", re.I)


@dataclass
class ColumnBinding:
    index: int
    header: str
    semantic: str
    header_score: int
    profile_score: float = 0.0
    explicit: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "index": self.index,
            "letter": _column_letter(self.index),
            "header": self.header,
            "semantic": self.semantic,
            "header_score": self.header_score,
            "profile_score": round(self.profile_score, 3),
            "explicit": self.explicit,
        }


@dataclass
class SheetSchema:
    sheet: str
    status: str
    confidence: float
    header_row: Optional[int] = None
    header_span: int = 1
    title_columns: List[ColumnBinding] = field(default_factory=list)
    venue_column: Optional[ColumnBinding] = None
    source_columns: List[ColumnBinding] = field(default_factory=list)
    year_column: Optional[ColumnBinding] = None
    reasons: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    candidate_columns: Dict[str, List[Dict[str, Any]]] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheet": self.sheet,
            "status": self.status,
            "confidence": round(self.confidence, 3),
            "header_row": self.header_row,
            "header_span": self.header_span,
            "columns": {
                "title": [column.to_dict() for column in self.title_columns],
                "venue": self.venue_column.to_dict() if self.venue_column else None,
                "sources": [column.to_dict() for column in self.source_columns],
                "year": self.year_column.to_dict() if self.year_column else None,
            },
            "reasons": self.reasons,
            "warnings": self.warnings,
            "candidate_columns": self.candidate_columns,
        }


@dataclass
class SchemaOptions:
    header_row: Optional[int] = None
    min_confidence: float = 0.72
    include_hidden_sheets: bool = False
    strict_schema: bool = False
    title_column: Optional[str] = None
    venue_column: Optional[str] = None
    source_columns: Optional[List[str]] = None
    schema_config: Optional[Mapping[str, Any]] = None


def _column_letter(index: int) -> str:
    value = index
    letters = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    text = re.sub(r"[\s_\-—–:：/\\|()（）\[\]【】]+", " ", text)
    return text.strip()


def compact_header(value: object) -> str:
    return re.sub(r"[^0-9a-z\u3400-\u9fff]+", "", normalize_header(value))


def header_score(value: object, aliases: Sequence[str]) -> int:
    normalized = normalize_header(value)
    compact = compact_header(value)
    if not compact:
        return 0
    best = 0
    for alias in aliases:
        alias_normalized = normalize_header(alias)
        alias_compact = compact_header(alias)
        if compact == alias_compact:
            best = max(best, 100)
        elif alias_compact and alias_compact in compact:
            best = max(best, 72 + min(len(alias_compact), 20))
        elif compact and compact in alias_compact:
            best = max(best, 48 + min(len(compact), 15))
        else:
            value_tokens = set(normalized.split())
            alias_tokens = set(alias_normalized.split())
            if value_tokens and alias_tokens:
                overlap = len(value_tokens & alias_tokens) / len(alias_tokens)
                if overlap >= 0.66:
                    best = max(best, int(45 + 35 * overlap))
    return min(best, 100)


def merge_aliases(config: Optional[Mapping[str, Any]]) -> Dict[str, List[str]]:
    aliases = {key: list(values) for key, values in DEFAULT_ALIASES.items()}
    if not config:
        return aliases
    supplied = config.get("aliases", {})
    if not isinstance(supplied, Mapping):
        raise ValueError("schema config 'aliases' must be an object")
    for semantic, values in supplied.items():
        if semantic not in aliases:
            raise ValueError(f"unknown alias semantic: {semantic}")
        if not isinstance(values, list) or not all(isinstance(v, str) for v in values):
            raise ValueError(f"aliases.{semantic} must be a string array")
        for value in values:
            if value not in aliases[semantic]:
                aliases[semantic].append(value)
    return aliases


def load_schema_config(path: Optional[Path]) -> Dict[str, Any]:
    if path is None:
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"cannot read schema config {path}: {exc}") from exc
    if not isinstance(data, dict):
        raise ValueError("schema config root must be a JSON object")
    return data


def _clean_url(url: str) -> str:
    value = unicodedata.normalize("NFKC", url).strip().rstrip(".,;，；。)]}〉》")
    return value


def extract_urls(
    value: object,
    hyperlink_target: Optional[str] = None,
    source_kind: str = "source",
) -> List[str]:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    found: List[str] = []
    if hyperlink_target:
        found.append(hyperlink_target.strip())
    if text.startswith("="):
        match = HYPERLINK_RE.search(text)
        if match:
            found.append(match.group(1))
    found.extend(URL_RE.findall(text))
    for doi in DOI_RE.findall(text):
        found.append(f"https://doi.org/{doi}")
    if source_kind == "arxiv" and text and not found:
        match = ARXIV_ID_RE.fullmatch(text.replace(" ", "")) or ARXIV_ID_RE.search(text)
        if match:
            found.append(f"https://arxiv.org/abs/{match.group(1)}")
    if source_kind == "openreview" and text and not found:
        simple_id = re.fullmatch(r"[A-Za-z0-9_-]{6,}", text)
        if simple_id:
            found.append(f"https://openreview.net/forum?id={text}")
    cleaned: List[str] = []
    for url in found:
        url = _clean_url(url)
        if not url:
            continue
        if url.startswith("www."):
            url = "https://" + url
        if url not in cleaned:
            cleaned.append(url)
    return cleaned


def normalize_doi(value: str) -> str:
    text = value.strip()
    if text.lower().startswith("https://doi.org/"):
        text = text.split("doi.org/", 1)[1]
    match = DOI_RE.search(text)
    return match.group(1).rstrip(".,;，；。)").casefold() if match else ""


def extract_arxiv_id(value: str) -> str:
    parts = urlsplit(value)
    if "arxiv.org" in (parts.hostname or "").casefold():
        match = re.search(r"/(?:abs|pdf)/([^/?#]+)", parts.path)
        if match:
            return match.group(1).removesuffix(".pdf").casefold()
    match = ARXIV_ID_RE.search(value)
    return match.group(1).casefold() if match else ""


def _merged_value(sheet: Any, row: int, column: int) -> object:
    cell = sheet.cell(row=row, column=column)
    if cell.value not in (None, ""):
        return cell.value
    coordinate = cell.coordinate
    for merged in sheet.merged_cells.ranges:
        if coordinate in merged:
            return sheet.cell(row=merged.min_row, column=merged.min_col).value
    return cell.value


def _header_labels(sheet: Any, end_row: int, span: int) -> List[str]:
    start_row = max(1, end_row - span + 1)
    labels: List[str] = []
    for column in range(1, sheet.max_column + 1):
        parts: List[str] = []
        for row in range(start_row, end_row + 1):
            value = _merged_value(sheet, row, column)
            text = unicodedata.normalize("NFKC", str(value or "")).strip()
            if text and text not in parts:
                parts.append(text)
        labels.append(" / ".join(parts))
    return labels


def _nonempty_values(sheet: Any, column: int, start_row: int, limit: int = PROFILE_ROWS) -> List[Tuple[object, Optional[str]]]:
    values: List[Tuple[object, Optional[str]]] = []
    end = min(sheet.max_row, start_row + limit - 1)
    for row in range(start_row, end + 1):
        cell = sheet.cell(row=row, column=column)
        if cell.value in (None, "") and not cell.hyperlink:
            continue
        values.append((cell.value, cell.hyperlink.target if cell.hyperlink else None))
    return values


def _source_profile(sheet: Any, column: int, start_row: int, source_kind: str) -> float:
    values = _nonempty_values(sheet, column, start_row)
    if not values:
        return 0.0
    matches = sum(bool(extract_urls(value, hyperlink, source_kind)) for value, hyperlink in values)
    return matches / len(values)


def _title_profile(sheet: Any, column: int, start_row: int) -> float:
    values = _nonempty_values(sheet, column, start_row)
    if not values:
        return 0.0
    good = 0
    unique: set[str] = set()
    for value, _ in values:
        text = unicodedata.normalize("NFKC", str(value or "")).strip()
        compact = re.sub(r"\s+", " ", text)
        if 5 <= len(compact) <= 500 and not URL_RE.search(compact) and not compact.isdigit():
            alpha = len(re.findall(r"[A-Za-z\u3400-\u9fff]", compact))
            if alpha >= max(3, int(len(compact) * 0.25)):
                good += 1
                unique.add(compact.casefold())
    ratio = good / len(values)
    uniqueness = len(unique) / max(good, 1)
    return min(1.0, ratio * 0.8 + uniqueness * 0.2)


def _sheet_has_data(sheet: Any) -> bool:
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value in (None, ""):
        return False
    checked = 0
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True):
        for value in row:
            checked += 1
            if value not in (None, ""):
                return True
            if checked >= 300:
                return False
    return False


def _semantic_scores(header: str, aliases: Mapping[str, Sequence[str]]) -> Dict[str, int]:
    return {semantic: header_score(header, aliases[semantic]) for semantic in aliases}


def _top_candidates(
    headers: Sequence[str],
    aliases: Mapping[str, Sequence[str]],
    semantic: str,
    limit: int = 5,
) -> List[Dict[str, Any]]:
    values = []
    for index, header in enumerate(headers, 1):
        score = header_score(header, aliases[semantic])
        if score:
            values.append({"index": index, "letter": _column_letter(index), "header": header, "score": score})
    return sorted(values, key=lambda item: (-item["score"], item["index"]))[:limit]


def _merged_candidates(
    headers: Sequence[str],
    aliases: Mapping[str, Sequence[str]],
    semantics: Sequence[str],
    limit: int = 8,
) -> List[Dict[str, Any]]:
    merged: Dict[int, Dict[str, Any]] = {}
    for semantic in semantics:
        for candidate in _top_candidates(headers, aliases, semantic, limit=len(headers)):
            index = int(candidate["index"])
            existing = merged.get(index)
            score = int(candidate["score"])
            if existing is None or score > int(existing["score"]):
                merged[index] = {**candidate, "semantic": semantic}
            elif score == int(existing["score"]):
                semantics_seen = set(str(existing.get("semantic", "")).split("|"))
                semantics_seen.add(semantic)
                existing["semantic"] = "|".join(sorted(value for value in semantics_seen if value))
    return sorted(merged.values(), key=lambda item: (-int(item["score"]), int(item["index"])))[:limit]


def _resolve_column_ref(ref: Any, headers: Sequence[str], label: str) -> int:
    if isinstance(ref, int):
        if 1 <= ref <= len(headers):
            return ref
        raise ValueError(f"{label} column index {ref} is outside 1..{len(headers)}")
    if not isinstance(ref, str) or not ref.strip():
        raise ValueError(f"{label} column reference must be a non-empty string or integer")
    text = ref.strip()
    if text.isdigit():
        return _resolve_column_ref(int(text), headers, label)
    if re.fullmatch(r"[A-Za-z]{1,3}", text):
        index = column_index_from_string(text.upper())
        return _resolve_column_ref(index, headers, label)
    normalized = compact_header(text)
    exact = [i for i, header in enumerate(headers, 1) if compact_header(header) == normalized]
    if len(exact) == 1:
        return exact[0]
    contains = [i for i, header in enumerate(headers, 1) if normalized and normalized in compact_header(header)]
    if len(contains) == 1:
        return contains[0]
    if len(exact) > 1 or len(contains) > 1:
        raise ValueError(f"{label} column reference {ref!r} is ambiguous")
    raise ValueError(f"cannot find {label} column {ref!r} in headers: {list(headers)}")


def _config_layers(config: Mapping[str, Any], sheet_name: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    global_mapping = config.get("global", {}) or {}
    if not isinstance(global_mapping, Mapping):
        raise ValueError("schema config 'global' must be an object")
    sheets = config.get("sheets", {}) or {}
    if not isinstance(sheets, Mapping):
        raise ValueError("schema config 'sheets' must be an object")
    specific = sheets.get(sheet_name, {}) or {}
    if not isinstance(specific, Mapping):
        raise ValueError(f"schema config sheets.{sheet_name} must be an object")
    return dict(global_mapping), dict(specific)


def _explicit_schema(
    sheet: Any,
    aliases: Mapping[str, Sequence[str]],
    options: SchemaOptions,
) -> Optional[SheetSchema]:
    config = dict(options.schema_config or {})
    global_mapping, specific_mapping = _config_layers(config, sheet.title)
    global_columns = global_mapping.get("columns", {}) or {}
    specific_columns = specific_mapping.get("columns", {}) or {}
    if not isinstance(global_columns, Mapping):
        raise ValueError("schema config global.columns must be an object")
    if not isinstance(specific_columns, Mapping):
        raise ValueError(f"sheet {sheet.title!r}: columns must be an object")

    def choose(column_key: str, cli_value: Any = None) -> Any:
        if column_key in specific_columns:
            return specific_columns[column_key]
        if cli_value:
            return cli_value
        return global_columns.get(column_key)

    title_ref = choose("title", options.title_column)
    title_en_ref = choose("title_en")
    title_zh_ref = choose("title_zh")
    source_refs: Any = choose("sources", options.source_columns) or choose("source", options.source_columns)
    venue_ref = choose("venue", options.venue_column)
    year_ref = choose("year")
    has_explicit_columns = any([title_ref, title_en_ref, title_zh_ref, source_refs, venue_ref, year_ref])
    if not has_explicit_columns:
        return None

    configured_header_row = specific_mapping.get(
        "header_row",
        options.header_row if options.header_row is not None else global_mapping.get("header_row", 1),
    )
    configured_header_span = specific_mapping.get("header_span", global_mapping.get("header_span", 1))
    header_row = int(configured_header_row)
    header_span = int(configured_header_span)
    if header_row < 1 or header_row > sheet.max_row:
        raise ValueError(f"sheet {sheet.title!r}: explicit header_row {header_row} is invalid")
    if not 1 <= header_span <= MAX_HEADER_SPAN:
        raise ValueError(f"sheet {sheet.title!r}: header_span must be 1..{MAX_HEADER_SPAN}")
    headers = _header_labels(sheet, header_row, header_span)

    if not title_ref and not title_en_ref and not title_zh_ref:
        raise ValueError(f"sheet {sheet.title!r}: explicit mapping requires a title column")
    if not source_refs:
        raise ValueError(f"sheet {sheet.title!r}: explicit mapping requires at least one source column")
    if not isinstance(source_refs, list):
        source_refs = [source_refs]

    title_columns: List[ColumnBinding] = []
    for semantic, ref in (("title_en", title_en_ref), ("title", title_ref), ("title_zh", title_zh_ref)):
        if ref:
            index = _resolve_column_ref(ref, headers, semantic)
            title_columns.append(ColumnBinding(index, headers[index - 1], semantic, 100, 1.0, True))

    global_kinds = global_columns.get("source_kinds", {}) or {}
    specific_kinds = specific_columns.get("source_kinds", {}) or {}
    if not isinstance(global_kinds, Mapping) or not isinstance(specific_kinds, Mapping):
        raise ValueError(f"sheet {sheet.title!r}: source_kinds must be an object")
    source_kinds = dict(global_kinds)
    source_kinds.update(specific_kinds)
    source_columns: List[ColumnBinding] = []
    for ref in source_refs:
        index = _resolve_column_ref(ref, headers, "source")
        kind = str(source_kinds.get(str(ref), source_kinds.get(_column_letter(index), "source")))
        if kind not in SOURCE_SEMANTICS:
            raise ValueError(f"sheet {sheet.title!r}: invalid source kind {kind!r}")
        source_columns.append(ColumnBinding(index, headers[index - 1], kind, 100, 1.0, True))

    venue_column = None
    if venue_ref:
        index = _resolve_column_ref(venue_ref, headers, "venue")
        venue_column = ColumnBinding(index, headers[index - 1], "venue", 100, 1.0, True)
    year_column = None
    if year_ref:
        index = _resolve_column_ref(year_ref, headers, "year")
        year_column = ColumnBinding(index, headers[index - 1], "year", 100, 1.0, True)

    return SheetSchema(
        sheet=sheet.title,
        status="accepted",
        confidence=1.0,
        header_row=header_row,
        header_span=header_span,
        title_columns=title_columns,
        venue_column=venue_column,
        source_columns=source_columns,
        year_column=year_column,
        reasons=["explicit column mapping"],
    )


def _auto_schema(
    sheet: Any,
    aliases: Mapping[str, Sequence[str]],
    options: SchemaOptions,
) -> SheetSchema:
    if not _sheet_has_data(sheet):
        return SheetSchema(sheet.title, "empty", 0.0, reasons=["sheet contains no data"])

    global_mapping, specific_mapping = _config_layers(dict(options.schema_config or {}), sheet.title)
    configured_header_row = specific_mapping.get(
        "header_row",
        options.header_row if options.header_row is not None else global_mapping.get("header_row"),
    )
    configured_header_span = specific_mapping.get("header_span", global_mapping.get("header_span"))

    row_candidates: Iterable[int]
    if configured_header_row is not None:
        row_candidates = [int(configured_header_row)]
    else:
        row_candidates = range(1, min(sheet.max_row, HEADER_SCAN_LIMIT) + 1)

    best: Optional[Tuple[float, SheetSchema]] = None
    profile_fallback: Optional[Tuple[float, SheetSchema]] = None
    for header_row in row_candidates:
        if header_row < 1 or header_row > sheet.max_row:
            continue
        span_candidates: Iterable[int]
        if configured_header_span is not None:
            span_candidates = [int(configured_header_span)]
        else:
            span_candidates = range(1, min(MAX_HEADER_SPAN, header_row) + 1)
        for span in span_candidates:
            if not 1 <= span <= min(MAX_HEADER_SPAN, header_row):
                continue
            headers = _header_labels(sheet, header_row, span)
            all_scores = [_semantic_scores(header, aliases) for header in headers]

            title_ranked: List[Tuple[int, int, str]] = []
            for index, scores in enumerate(all_scores, 1):
                semantic = max(TITLE_SEMANTICS, key=lambda key: scores[key])
                title_ranked.append((scores[semantic], index, semantic))
            title_ranked.sort(key=lambda value: (-value[0], value[1], value[2]))
            title_score, title_index, title_semantic = title_ranked[0] if title_ranked else (0, 0, "title")

            source_bindings: List[ColumnBinding] = []
            for index, scores in enumerate(all_scores, 1):
                semantic = max(SOURCE_SEMANTICS, key=lambda key: scores[key])
                score = scores[semantic]
                profile = _source_profile(sheet, index, header_row + 1, semantic)
                if score >= 52 or (score >= 35 and profile >= 0.65):
                    source_bindings.append(ColumnBinding(index, headers[index - 1], semantic, score, profile))
            source_bindings.sort(key=lambda col: (-col.header_score, -col.profile_score, col.index))

            if title_score <= 0 or not source_bindings:
                # Profile-only inference is never accepted automatically. It exists so
                # an Agent can see likely columns and create an explicit mapping rather
                # than guessing or lowering the safety threshold.
                profile_title_candidates = [
                    (_title_profile(sheet, index, header_row + 1), index)
                    for index in range(1, sheet.max_column + 1)
                ]
                profile_source_candidates = [
                    (_source_profile(sheet, index, header_row + 1, "source"), index)
                    for index in range(1, sheet.max_column + 1)
                ]
                best_title_profile, fallback_title_index = max(profile_title_candidates, default=(0.0, 0))
                best_source_profile, fallback_source_index = max(profile_source_candidates, default=(0.0, 0))
                if (
                    fallback_title_index
                    and fallback_source_index
                    and fallback_title_index != fallback_source_index
                    and best_title_profile >= 0.70
                    and best_source_profile >= 0.60
                ):
                    fallback_confidence = min(0.59, 0.30 * best_title_profile + 0.25 * best_source_profile)
                    fallback_schema = SheetSchema(
                        sheet=sheet.title,
                        status="low_confidence",
                        confidence=fallback_confidence,
                        header_row=header_row,
                        header_span=span,
                        title_columns=[
                            ColumnBinding(
                                fallback_title_index,
                                headers[fallback_title_index - 1],
                                "title",
                                0,
                                best_title_profile,
                            )
                        ],
                        source_columns=[
                            ColumnBinding(
                                fallback_source_index,
                                headers[fallback_source_index - 1],
                                "source",
                                0,
                                best_source_profile,
                            )
                        ],
                        reasons=[
                            "profile-only inference; header semantics were not recognized",
                            f"likely title profile={best_title_profile:.2f}",
                            f"likely source profile={best_source_profile:.2f}",
                        ],
                        warnings=[
                            "do not lower the confidence threshold; confirm these columns with --title-column and --source-columns or --schema-config"
                        ],
                        candidate_columns={
                            "title": [
                                {
                                    "index": fallback_title_index,
                                    "letter": _column_letter(fallback_title_index),
                                    "header": headers[fallback_title_index - 1],
                                    "score": 0,
                                    "profile_score": round(best_title_profile, 3),
                                }
                            ],
                            "source": [
                                {
                                    "index": fallback_source_index,
                                    "letter": _column_letter(fallback_source_index),
                                    "header": headers[fallback_source_index - 1],
                                    "score": 0,
                                    "profile_score": round(best_source_profile, 3),
                                }
                            ],
                            "venue": [],
                        },
                    )
                    fallback_rank = fallback_confidence - 0.002 * header_row - 0.01 * (span - 1)
                    if profile_fallback is None or fallback_rank > profile_fallback[0]:
                        profile_fallback = (fallback_rank, fallback_schema)
                continue

            title_profile = _title_profile(sheet, title_index, header_row + 1)
            source_header = max(col.header_score for col in source_bindings) / 100
            source_profile = max(col.profile_score for col in source_bindings)
            title_conf = 0.84 * (title_score / 100) + 0.16 * title_profile
            source_conf = 0.80 * source_header + 0.20 * source_profile
            ambiguity_penalty = 0.0
            if len(title_ranked) > 1 and title_ranked[1][0] >= max(50, title_score - 8):
                ambiguity_penalty += 0.08
            if title_index in {column.index for column in source_bindings}:
                ambiguity_penalty += 0.25
            span_penalty = 0.015 * (span - 1)
            confidence = max(0.0, min(1.0, 0.54 * title_conf + 0.46 * source_conf - ambiguity_penalty - span_penalty))

            title_columns = [
                ColumnBinding(title_index, headers[title_index - 1], title_semantic, title_score, title_profile)
            ]
            # Add separately labelled EN/ZH title columns when present.
            for semantic in ("title_en", "title_zh"):
                candidates = [
                    (scores[semantic], index)
                    for index, scores in enumerate(all_scores, 1)
                    if index != title_index
                ]
                score, index = min(candidates, key=lambda value: (-value[0], value[1]), default=(0, 0))
                if score >= 70:
                    title_columns.append(
                        ColumnBinding(index, headers[index - 1], semantic, score, _title_profile(sheet, index, header_row + 1))
                    )

            venue_candidates = [(scores["venue"], index) for index, scores in enumerate(all_scores, 1)]
            venue_score, venue_index = min(venue_candidates, key=lambda value: (-value[0], value[1]), default=(0, 0))
            venue_column = (
                ColumnBinding(venue_index, headers[venue_index - 1], "venue", venue_score)
                if venue_score >= 55
                else None
            )
            year_candidates = [(scores["year"], index) for index, scores in enumerate(all_scores, 1)]
            year_score, year_index = min(year_candidates, key=lambda value: (-value[0], value[1]), default=(0, 0))
            year_column = (
                ColumnBinding(year_index, headers[year_index - 1], "year", year_score)
                if year_score >= 60
                else None
            )

            warnings: List[str] = []
            if venue_column is None:
                warnings.append("venue column not detected; records will use 未分类")
            if title_score < 70:
                warnings.append("title column match is weak; consider an explicit mapping")
            if max(column.header_score for column in source_bindings) < 70:
                warnings.append("source column match is weak; consider an explicit mapping")
            if len(source_bindings) > 1:
                warnings.append(f"{len(source_bindings)} source columns detected and will be merged by priority")

            schema = SheetSchema(
                sheet=sheet.title,
                status="accepted" if confidence >= options.min_confidence else "low_confidence",
                confidence=confidence,
                header_row=header_row,
                header_span=span,
                title_columns=title_columns,
                venue_column=venue_column,
                source_columns=source_bindings,
                year_column=year_column,
                reasons=[
                    f"title header score={title_score}",
                    f"best source header score={max(column.header_score for column in source_bindings)}",
                    f"title profile={title_profile:.2f}",
                    f"source profile={source_profile:.2f}",
                ],
                warnings=warnings,
                candidate_columns={
                    "title": _merged_candidates(headers, aliases, TITLE_SEMANTICS),
                    "source": _merged_candidates(headers, aliases, SOURCE_SEMANTICS),
                    "venue": _merged_candidates(headers, aliases, ("venue",)),
                },
            )
            rank = confidence - 0.0005 * (header_row - 1) - span_penalty
            if best is None or rank > best[0]:
                best = (rank, schema)

    if best is None:
        if profile_fallback is not None:
            return profile_fallback[1]
        return SheetSchema(
            sheet.title,
            "unrecognized",
            0.0,
            reasons=["no row contains both a recognizable title field and a source/DOI field"],
        )
    return best[1]


def detect_sheet_schema(
    sheet: Any,
    options: SchemaOptions,
    aliases: Optional[Mapping[str, Sequence[str]]] = None,
) -> SheetSchema:
    aliases = aliases or DEFAULT_ALIASES
    if sheet.sheet_state != "visible" and not options.include_hidden_sheets:
        return SheetSchema(sheet.title, "hidden", 0.0, reasons=[f"sheet state is {sheet.sheet_state}"])
    explicit = _explicit_schema(sheet, aliases, options)
    if explicit:
        return explicit
    return _auto_schema(sheet, aliases, options)


def title_from_row(sheet: Any, row: int, schema: SheetSchema) -> Tuple[str, str, str]:
    values: Dict[str, str] = {}
    for binding in schema.title_columns:
        text = unicodedata.normalize("NFKC", str(sheet.cell(row=row, column=binding.index).value or "")).strip()
        if text:
            values[binding.semantic] = text
    generic = values.get("title", "")
    title_en = values.get("title_en", "")
    title_zh = values.get("title_zh", "")
    if generic:
        match = re.search(
            r"\[\s*EN\s*\]\s*(.*?)(?=\[\s*(?:中文|CN|ZH)\s*\]|$)",
            generic,
            flags=re.I | re.S,
        )
        if match and not title_en:
            title_en = re.sub(r"\s+", " ", match.group(1)).strip()
        zh_match = re.search(r"\[\s*(?:中文|CN|ZH)\s*\]\s*(.*)$", generic, flags=re.I | re.S)
        if zh_match and not title_zh:
            title_zh = re.sub(r"\s+", " ", zh_match.group(1)).strip()
    generic_first = generic.splitlines()[0].strip() if generic.splitlines() else ""
    preferred = title_en or generic_first or title_zh
    preferred = re.sub(r"\s+", " ", preferred).strip()
    return preferred, re.sub(r"\s+", " ", title_en).strip(), re.sub(r"\s+", " ", title_zh).strip()


def source_values_from_row(sheet: Any, row: int, schema: SheetSchema) -> Tuple[List[str], List[Dict[str, Any]]]:
    urls: List[str] = []
    details: List[Dict[str, Any]] = []
    for binding in schema.source_columns:
        cell = sheet.cell(row=row, column=binding.index)
        extracted = extract_urls(
            cell.value,
            cell.hyperlink.target if cell.hyperlink else None,
            binding.semantic,
        )
        details.append(
            {
                "column": binding.index,
                "letter": _column_letter(binding.index),
                "header": binding.header,
                "semantic": binding.semantic,
                "raw": str(cell.value or ""),
                "urls": extracted,
            }
        )
        for url in extracted:
            if url not in urls:
                urls.append(url)
    return urls, details


def derive_title_from_sources(urls: Sequence[str], row: int) -> str:
    for url in urls:
        doi = normalize_doi(url)
        if doi:
            return "DOI_" + re.sub(r"[^0-9a-z._-]+", "_", doi, flags=re.I)
        arxiv = extract_arxiv_id(url)
        if arxiv:
            return "arXiv_" + arxiv
    for url in urls:
        parts = urlsplit(url)
        leaf = Path(parts.path).stem
        if leaf and leaf.casefold() not in {"download", "pdf", "view", "article"}:
            return re.sub(r"[^0-9A-Za-z._-]+", "_", leaf)[:120]
    return f"paper_row_{row}"


def dedupe_identity(title: str, urls: Sequence[str]) -> Tuple[str, str]:
    for url in urls:
        doi = normalize_doi(url)
        if doi:
            return f"doi:{doi}", "doi"
    for url in urls:
        arxiv = extract_arxiv_id(url)
        if arxiv:
            return f"arxiv:{arxiv}", "arxiv"
    normalized = unicodedata.normalize("NFKC", title).casefold()
    normalized = re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return f"title:{normalized}", "title"
