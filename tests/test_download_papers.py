from __future__ import annotations

import csv
import json
import sys
import tempfile
import threading
import unittest
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import download_papers as dp  # noqa: E402
import validate_collection as vc  # noqa: E402


def fake_pdf() -> bytes:
    body = b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog >>\nendobj\n"
    body += b"% test-data\n" * 100
    body += b"trailer\n<<>>\n%%EOF\n"
    return body


class TestTextParsing(unittest.TestCase):
    def test_english_title_bilingual_variants(self):
        self.assertEqual(
            dp.english_title("[EN] A Reliable Paper\n[中文] 一篇论文"),
            "A Reliable Paper",
        )
        self.assertEqual(
            dp.english_title("[EN] A Reliable Paper [中文] 一篇论文"),
            "A Reliable Paper",
        )
        self.assertEqual(dp.english_title("Plain title\nSecond line"), "Plain title")

    def test_clean_name_is_bounded_and_safe(self):
        value = dp.clean_name('A/B:C*D?E"F<G>H|I', limit=20)
        self.assertNotRegex(value, r'[/:*?"<>|]')
        self.assertLessEqual(len(value), 20)

    def test_extract_urls(self):
        self.assertEqual(
            dp.extract_urls('=HYPERLINK("https://example.org/a","paper")'),
            ["https://example.org/a"],
        )
        self.assertEqual(dp.extract_urls("10.1234/example"), ["https://doi.org/10.1234/example"])

    def test_candidate_urls(self):
        urls = dp.candidate_urls("https://arxiv.org/abs/2401.12345")
        self.assertEqual(urls[0], "https://arxiv.org/pdf/2401.12345.pdf")
        self.assertIn("https://arxiv.org/abs/2401.12345", urls)
        openreview = dp.candidate_urls("https://openreview.net/forum?id=abc123")
        self.assertEqual(openreview[0], "https://openreview.net/pdf?id=abc123")

    def test_html_pdf_parser_accepts_meta_and_relative_link(self):
        html = b'''<html><head><meta name="citation_pdf_url" content="/paper.pdf"></head>
        <body><a href="/other.pdf">PDF</a></body></html>'''
        links = dp.html_pdf_links("https://example.org/page", html)
        self.assertEqual(links[0], "https://example.org/paper.pdf")
        self.assertIn("https://example.org/other.pdf", links)

    def test_custom_header_parser_and_interactive_html_classification(self):
        headers = dp.parse_request_headers(["Authorization: Bearer abc", "X-Test: value:with:colon"])
        self.assertEqual(headers["Authorization"], "Bearer abc")
        self.assertEqual(headers["X-Test"], "value:with:colon")
        self.assertEqual(dp.classify_interactive_html(b"<html>Sign in to access this article</html>"), "session_required")
        self.assertEqual(dp.classify_interactive_html(b"<html><div class='hcaptcha'></div></html>"), "interactive_challenge")
        self.assertEqual(dp._classify_http_status(401), "authorization_required")
        self.assertEqual(dp._classify_http_status(403), "forbidden_or_challenge")


class TestWorkbookInspection(unittest.TestCase):
    def make_workbook(self, path: Path, header_row: int = 3) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "CCF-A论文"
        sheet.cell(header_row, 1, "Title / 论文标题")
        sheet.cell(header_row, 2, "会议/期刊")
        sheet.cell(header_row, 3, "URL")
        sheet.cell(header_row + 1, 1, "[EN] Paper One\n[中文] 论文一")
        sheet.cell(header_row + 1, 2, "CVPR")
        url_cell = sheet.cell(header_row + 1, 3, "paper")
        url_cell.hyperlink = "https://example.org/paper"
        workbook.save(path)

    def test_auto_detect_header_and_hyperlink(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "papers.xlsx"
            self.make_workbook(path, header_row=5)
            report, items = dp.inspect_workbook(path)
            self.assertTrue(report["ok"])
            self.assertEqual(report["records"], 1)
            self.assertEqual(report["worksheets"][0]["header_row"], 5)
            self.assertEqual(items[0]["title"], "Paper One")
            self.assertEqual(items[0]["source_urls"], ["https://example.org/paper"])

    def test_unknown_sheet_is_error(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "papers.xlsx"
            self.make_workbook(path)
            with self.assertRaises(ValueError):
                dp.inspect_workbook(path, ["missing"])

    def test_filename_collision_gets_hash_suffix(self):
        with tempfile.TemporaryDirectory() as temp:
            output = Path(temp)
            items = [
                {"sheet": "S", "venue": "V", "title": "A/B", "title_key": "first"},
                {"sheet": "S", "venue": "V", "title": "A:B", "title_key": "second"},
            ]
            dp.assign_targets(items, output)
            self.assertNotEqual(items[0]["target"], items[1]["target"])
            self.assertIn("__", Path(items[1]["target"]).stem)


class LocalHandler(BaseHTTPRequestHandler):
    pdf = fake_pdf()
    flaky_count = 0

    def do_GET(self):
        if self.path == "/landing":
            payload = b'<html><head><meta name="citation_pdf_url" content="/paper.pdf"></head></html>'
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return
        if self.path == "/cookie-landing":
            payload = b'<html><head><meta name="citation_pdf_url" content="/cookie-paper.pdf"></head></html>'
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.send_header("Set-Cookie", "paper_session=ok; Path=/")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return
        if self.path == "/cookie-paper.pdf":
            cookie = self.headers.get("Cookie", "")
            referer = self.headers.get("Referer", "")
            if "paper_session=ok" not in cookie or not referer.endswith("/cookie-landing"):
                self.send_response(403)
                self.end_headers()
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Length", str(len(self.pdf)))
            self.end_headers()
            self.wfile.write(self.pdf)
            return
        if self.path == "/paper.pdf":
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Length", str(len(self.pdf)))
            self.end_headers()
            self.wfile.write(self.pdf)
            return
        if self.path == "/auth-paper.pdf":
            authorization = self.headers.get("Authorization", "")
            cookie = self.headers.get("Cookie", "")
            if authorization != "Bearer test-token" or "licensed=yes" not in cookie:
                self.send_response(403)
                self.end_headers()
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Length", str(len(self.pdf)))
            self.end_headers()
            self.wfile.write(self.pdf)
            return
        if self.path == "/login-page":
            payload = b"<html><body>Sign in to access this article</body></html>"
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return
        self.send_response(404)
        self.end_headers()

    def log_message(self, format, *args):
        return


class TestNetworkAndValidation(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.server = ThreadingHTTPServer(("127.0.0.1", 0), LocalHandler)
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()
        cls.base = f"http://127.0.0.1:{cls.server.server_address[1]}"

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.server.server_close()

    def test_private_host_blocked_by_default(self):
        error = dp.url_safety_error(f"{self.base}/paper.pdf", False)
        self.assertIn("blocked", error)

    def test_fetch_follows_html_pdf_link(self):
        with tempfile.TemporaryDirectory() as temp:
            destination = Path(temp) / "paper.pdf"
            final_url, error = dp.fetch_pdf(
                [f"{self.base}/landing"],
                destination,
                timeout=3,
                retries=0,
                max_bytes=1024 * 1024,
                min_pdf_bytes=1024,
                allow_private_hosts=True,
            )
            self.assertEqual(error, "")
            self.assertTrue(final_url.endswith("/paper.pdf"))
            self.assertTrue(dp.valid_pdf(destination))


    def test_html_discovery_preserves_cookie_and_referer(self):
        with tempfile.TemporaryDirectory() as temp:
            destination = Path(temp) / "cookie-paper.pdf"
            result = dp.fetch_pdf_detailed(
                [f"{self.base}/cookie-landing"],
                destination,
                timeout=3,
                retries=0,
                max_bytes=1024 * 1024,
                min_pdf_bytes=1024,
                allow_private_hosts=True,
                host_delay=0,
            )
            self.assertEqual(result.category, "success", result.error)
            self.assertTrue(result.final_url.endswith("/cookie-paper.pdf"))
            self.assertTrue(dp.valid_pdf(destination))

    def test_user_supplied_authorization_and_cookie_are_applied(self):
        with tempfile.TemporaryDirectory() as temp:
            destination = Path(temp) / "auth-paper.pdf"
            result = dp.fetch_pdf_detailed(
                [f"{self.base}/auth-paper.pdf"],
                destination,
                timeout=3,
                retries=0,
                max_bytes=1024 * 1024,
                min_pdf_bytes=1024,
                allow_private_hosts=True,
                host_delay=0,
                request_headers={"Authorization": "Bearer test-token"},
                cookie_header="licensed=yes",
            )
            self.assertEqual(result.category, "success", result.error)
            self.assertTrue(dp.valid_pdf(destination))

    def test_netscape_cookie_file_is_supported(self):
        with tempfile.TemporaryDirectory() as temp:
            temp_path = Path(temp)
            destination = temp_path / "auth-paper.pdf"
            cookie_file = temp_path / "cookies.txt"
            cookie_file.write_text(
                "# Netscape HTTP Cookie File\n127.0.0.1\tFALSE\t/\tFALSE\t2147483647\tlicensed\tyes\n",
                encoding="utf-8",
            )
            result = dp.fetch_pdf_detailed(
                [f"{self.base}/auth-paper.pdf"],
                destination,
                timeout=3,
                retries=0,
                max_bytes=1024 * 1024,
                min_pdf_bytes=1024,
                allow_private_hosts=True,
                host_delay=0,
                request_headers={"Authorization": "Bearer test-token"},
                cookie_file=cookie_file,
            )
            self.assertEqual(result.category, "success", result.error)
            self.assertTrue(dp.valid_pdf(destination))

    def test_end_to_end_auth_context_is_used_but_secrets_are_not_logged(self):
        with tempfile.TemporaryDirectory() as temp:
            temp_path = Path(temp)
            workbook_path = temp_path / "auth.xlsx"
            output = temp_path / "collection"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Auth"
            sheet.append(["Title", "URL"])
            sheet.append(["Authenticated Paper", f"{self.base}/auth-paper.pdf"])
            workbook.save(workbook_path)

            args = dp.parse_args([
                "--input", str(workbook_path),
                "--output", str(output),
                "--allow-private-hosts",
                "--authorization", "Bearer test-token",
                "--cookie", "licensed=yes",
                "--host-delay", "0",
                "--retries", "0",
            ])
            summary = dp.run(args)
            self.assertEqual(summary["record_success"], 1)
            journal = (output / "运行日志.jsonl").read_text(encoding="utf-8")
            self.assertIn("Authorization", journal)
            self.assertIn('"cookie_supplied": true', journal)
            self.assertNotIn("test-token", journal)
            self.assertNotIn("licensed=yes", journal)

    def test_login_page_is_recoverable_category(self):
        with tempfile.TemporaryDirectory() as temp:
            destination = Path(temp) / "login.pdf"
            result = dp.fetch_pdf_detailed(
                [f"{self.base}/login-page"],
                destination,
                timeout=3,
                retries=0,
                max_bytes=1024 * 1024,
                min_pdf_bytes=1024,
                allow_private_hosts=True,
                host_delay=0,
            )
            self.assertEqual(result.category, "session_required")
            self.assertFalse(destination.exists())


    def test_force_failure_preserves_existing_valid_pdf(self):
        with tempfile.TemporaryDirectory() as temp:
            temp_path = Path(temp)
            workbook_path = temp_path / "papers.xlsx"
            output = temp_path / "collection"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "S"
            sheet.append(["Title", "会议/期刊", "URL"])
            sheet.append(["Existing Paper", "V", "http://127.0.0.1:1/missing.pdf"])
            workbook.save(workbook_path)

            existing = output / "S" / "V" / "Existing Paper.pdf"
            existing.parent.mkdir(parents=True)
            existing.write_bytes(fake_pdf())

            args = dp.parse_args(
                [
                    "--input",
                    str(workbook_path),
                    "--output",
                    str(output),
                    "--force",
                    "--timeout",
                    "1",
                    "--retries",
                    "0",
                ]
            )
            summary = dp.run(args)
            self.assertEqual(summary["record_success"], 1)
            with (output / "下载清单.csv").open(encoding="utf-8-sig", newline="") as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(row["resolution"], "reused-after-download-failure")
            self.assertTrue(dp.valid_pdf(existing))

    def test_end_to_end_run_and_validate(self):
        with tempfile.TemporaryDirectory() as temp:
            temp_path = Path(temp)
            workbook_path = temp_path / "papers.xlsx"
            output = temp_path / "collection"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "推荐精读"
            sheet.append(["Title", "会议/期刊", "URL"])
            sheet.append(["[EN] Test Paper\n[中文] 测试论文", "TestConf", f"{self.base}/landing"])
            workbook.save(workbook_path)

            args = dp.parse_args(
                [
                    "--input",
                    str(workbook_path),
                    "--output",
                    str(output),
                    "--allow-private-hosts",
                    "--timeout",
                    "3",
                ]
            )
            summary = dp.run(args)
            self.assertEqual(summary["record_success"], 1)
            manifest = output / "下载清单.csv"
            result = vc.validate(output, manifest, strict=False, min_pdf_bytes=1024)
            self.assertTrue(result["ok"], result)
            with manifest.open(encoding="utf-8-sig", newline="") as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(row["status"], "成功")
            self.assertEqual(len(row["sha256"]), 64)
            self.assertTrue((output / row["local_path"]).is_file())
            report = json.loads((output / "下载报告.json").read_text(encoding="utf-8"))
            self.assertEqual(report["unique_success"], 1)


    def test_resume_recovers_verified_interrupted_cache_without_network(self):
        with tempfile.TemporaryDirectory() as temp:
            temp_path = Path(temp)
            workbook_path = temp_path / "papers.xlsx"
            output = temp_path / "collection"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Title", "URL"])
            sheet.append(["Cached Paper", "http://127.0.0.1:1/unreachable.pdf"])
            workbook.save(workbook_path)

            report, items = dp.inspect_workbook(workbook_path)
            self.assertTrue(report["ok"], report)
            key = items[0]["dedupe_key"]
            cache_path = output / ".download-cache" / (dp.hashlib.sha256(key.encode("utf-8")).hexdigest()[:24] + ".pdf")
            cache_path.parent.mkdir(parents=True)
            cache_path.write_bytes(fake_pdf())

            args = dp.parse_args(
                [
                    "--input", str(workbook_path),
                    "--output", str(output),
                    "--resume",
                    "--timeout", "1",
                    "--retries", "0",
                ]
            )
            summary = dp.run(args)
            self.assertEqual(summary["record_success"], 1)
            with (output / "下载清单.csv").open(encoding="utf-8-sig", newline="") as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(row["resolution"], "recovered-cache")
            self.assertEqual(json.loads(row["attempts"]), [])


class TestAdaptiveSchemas(unittest.TestCase):
    def test_multirow_merged_header_and_multiple_source_columns(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "multirow.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "论文库"
            sheet.merge_cells("A1:A2")
            sheet["A1"] = "论文标题"
            sheet.merge_cells("B1:C1")
            sheet["B1"] = "公开来源"
            sheet["B2"] = "PDF链接"
            sheet["C2"] = "DOI"
            sheet.merge_cells("D1:D2")
            sheet["D1"] = "期刊/会议"
            sheet.append(["Adaptive Paper", "https://example.org/paper.pdf", "10.1234/adaptive", "TGRS"])
            workbook.save(path)

            report, items = dp.inspect_workbook(path)
            self.assertTrue(report["ok"], report)
            schema = report["worksheets"][0]
            self.assertEqual(schema["header_row"], 2)
            self.assertIn(schema["header_span"], {1, 2})
            self.assertEqual(len(schema["columns"]["sources"]), 2)
            self.assertEqual(items[0]["venue"], "TGRS")
            self.assertIn("https://example.org/paper.pdf", items[0]["source_urls"])
            self.assertIn("https://doi.org/10.1234/adaptive", items[0]["source_urls"])

    def test_irrelevant_summary_sheet_is_skipped(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "mixed.xlsx"
            workbook = openpyxl.Workbook()
            summary = workbook.active
            summary.title = "说明"
            summary.append(["本工作簿用于论文整理"])
            papers = workbook.create_sheet("论文")
            papers.append(["文献名称", "原文地址"])
            papers.append(["Paper", "https://example.org/paper.pdf"])
            workbook.save(path)

            report, items = dp.inspect_workbook(path)
            self.assertTrue(report["ok"], report)
            self.assertEqual(report["accepted_sheets"], ["论文"])
            self.assertEqual(len(items), 1)
            skipped = next(sheet for sheet in report["worksheets"] if sheet["sheet"] == "说明")
            self.assertIn(skipped["status"], {"unrecognized", "low_confidence"})

    def test_selected_unrecognized_sheet_is_error(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "unknown.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "说明"
            sheet.append(["内容", "备注"])
            sheet.append(["abc", "def"])
            workbook.save(path)

            report, _ = dp.inspect_workbook(path, ["说明"])
            self.assertFalse(report["ok"])
            self.assertTrue(any(issue["severity"] == "error" for issue in report["issues"]))

    def test_explicit_mapping_handles_unknown_headers(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "custom.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "自定义"
            sheet.append(["作品名", "资源入口", "载体"])
            sheet.append(["Mapped Paper", "10.5555/mapped", "Mapped Journal"])
            workbook.save(path)
            config = {
                "sheets": {
                    "自定义": {
                        "header_row": 1,
                        "columns": {
                            "title": "作品名",
                            "sources": ["资源入口"],
                            "venue": "载体",
                            "source_kinds": {"资源入口": "doi"},
                        },
                    }
                }
            }
            report, items = dp.inspect_workbook(path, schema_config=config)
            self.assertTrue(report["ok"], report)
            self.assertEqual(report["worksheets"][0]["confidence"], 1.0)
            self.assertEqual(items[0]["source_urls"], ["https://doi.org/10.5555/mapped"])

    def test_custom_aliases_enable_auto_detection(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "aliases.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["作品名", "公开获取"])
            sheet.append(["Alias Paper", "https://example.org/a.pdf"])
            workbook.save(path)
            config = {"aliases": {"title": ["作品名"], "pdf_url": ["公开获取"]}}
            report, items = dp.inspect_workbook(path, schema_config=config)
            self.assertTrue(report["ok"], report)
            self.assertEqual(len(items), 1)

    def test_doi_is_preferred_for_deduplication(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "dedupe.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Title", "DOI"])
            sheet.append(["Paper Version A", "10.1000/same"])
            sheet.append(["Paper Version B", "https://doi.org/10.1000/same"])
            workbook.save(path)
            report, items = dp.inspect_workbook(path)
            self.assertTrue(report["ok"], report)
            self.assertEqual(report["unique_papers"], 1)
            self.assertEqual(items[0]["dedupe_key"], items[1]["dedupe_key"])
            self.assertEqual(items[0]["dedupe_mode"], "doi")

    def test_missing_title_is_deterministically_derived(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "derived.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Title", "DOI"])
            sheet.append(["", "10.1000/derived"])
            workbook.save(path)
            report, items = dp.inspect_workbook(path)
            self.assertTrue(report["ok"], report)
            self.assertEqual(items[0]["title"], "DOI_10.1000_derived")
            self.assertTrue(items[0]["title_derived"])

    def test_low_confidence_profile_does_not_blindly_download(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "ambiguous.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["字段甲", "字段乙"])
            sheet.append(["A plausible long paper name", "https://example.org/a.pdf"])
            sheet.append(["Another plausible paper name", "https://example.org/b.pdf"])
            workbook.save(path)
            report, items = dp.inspect_workbook(path)
            self.assertFalse(report["ok"])
            self.assertEqual(items, [])
            self.assertEqual(report["accepted_sheets"], [])


    def test_global_header_only_config_keeps_auto_detection(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "global-header.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["说明", ""])
            sheet.append(["Title", "URL"])
            sheet.append(["Configured Header Paper", "https://example.org/paper.pdf"])
            workbook.save(path)

            report, items = dp.inspect_workbook(
                path,
                schema_config={"global": {"header_row": 2, "header_span": 1}},
            )
            self.assertTrue(report["ok"], report)
            self.assertEqual(report["worksheets"][0]["header_row"], 2)
            self.assertFalse(report["worksheets"][0]["columns"]["title"][0]["explicit"])
            self.assertEqual(items[0]["title"], "Configured Header Paper")

    def test_per_sheet_mapping_overrides_cli_and_global_columns(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "precedence.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Special"
            sheet.append(["Wrong title", "Correct title", "Wrong source", "Correct source"])
            sheet.append(["Do not use", "Preferred Paper", "https://example.org/wrong.pdf", "10.7777/preferred"])
            workbook.save(path)
            config = {
                "global": {
                    "columns": {"title": "Wrong title", "sources": ["Wrong source"]}
                },
                "sheets": {
                    "Special": {
                        "columns": {
                            "title": "Correct title",
                            "sources": ["Correct source"],
                            "source_kinds": {"Correct source": "doi"},
                        }
                    }
                },
            }
            report, items = dp.inspect_workbook(
                path,
                schema_config=config,
                title_column="A",
                source_columns=["C"],
            )
            self.assertTrue(report["ok"], report)
            self.assertEqual(items[0]["title"], "Preferred Paper")
            self.assertEqual(items[0]["source_urls"], ["https://doi.org/10.7777/preferred"])

    def test_profile_only_report_exposes_candidate_columns(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "profile-candidates.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["字段甲", "字段乙"])
            sheet.append(["A sufficiently descriptive paper title", "https://example.org/a.pdf"])
            sheet.append(["Another sufficiently descriptive title", "https://example.org/b.pdf"])
            workbook.save(path)

            report, _ = dp.inspect_workbook(path)
            schema = report["worksheets"][0]
            self.assertEqual(schema["status"], "low_confidence")
            self.assertEqual(schema["candidate_columns"]["title"][0]["letter"], "A")
            self.assertEqual(schema["candidate_columns"]["source"][0]["letter"], "B")

    def test_pdf_semantic_boost_prioritizes_download_endpoint(self):
        details = [
            {"semantic": "doi", "urls": ["https://doi.org/10.1000/example"]},
            {"semantic": "pdf_url", "urls": ["https://example.org/download?id=123"]},
        ]
        ranked = dp.rank_source_details(details)
        self.assertEqual(ranked[0], "https://example.org/download?id=123")

    def test_source_ranking_ignores_repository_before_pdf(self):
        ranked = dp.rank_source_urls(
            ["https://github.com/example/project", "https://example.org/paper.pdf", "https://doi.org/10.1000/x"]
        )
        self.assertEqual(ranked[0], "https://example.org/paper.pdf")
        self.assertEqual(ranked[-1], "https://github.com/example/project")


if __name__ == "__main__":
    unittest.main()
